import json
import os
import shutil
import tempfile

import openpyxl

import app as app_module
import history_store
import read_seed


def _build_fixture(folder):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A9"] = "업체명"
    ws["D9"] = "테스트업체"
    ws["L11"] = "청 구 인"
    ws["O11"] = "테스트연구원"
    ws["A12"] = "내  용"
    ws["D12"] = "테스트 구매의 건"
    ws["A13"] = "상세내용"
    ws["B14"] = "테스트 상세내용입니다."
    ws["B16"] = "중앙행정기관 : 테스트부처"
    ws["B17"] = "전문기관 : 테스트기관"
    ws["B18"] = "과제명 : 테스트 과제명"
    ws["B20"] = "연구재료비 집행의 건 (연구개발계획서 상 계상되어 있는 건임)"
    wb.save(os.path.join(folder, "fixture.xlsx"))


def test_refresh_read_seed_route():
    read_dir = tempfile.mkdtemp()
    data_dir = tempfile.mkdtemp()
    original_read_dir = read_seed.READ_DIR
    original_history_path = history_store.HISTORY_PATH
    original_projects_path = history_store.PROJECTS_PATH
    try:
        _build_fixture(read_dir)
        read_seed.READ_DIR = read_dir
        history_store.HISTORY_PATH = os.path.join(data_dir, "history.json")
        history_store.PROJECTS_PATH = os.path.join(data_dir, "projects.json")

        client = app_module.app.test_client()
        resp = client.post("/refresh_read_seed")
        assert resp.status_code == 302
        assert "/?refreshed=1&added=" in resp.headers["Location"]

        with open(history_store.HISTORY_PATH, "r", encoding="utf-8") as f:
            saved_history = json.load(f)
        assert "테스트업체" in saved_history["company"]

        with open(history_store.PROJECTS_PATH, "r", encoding="utf-8") as f:
            saved_projects = json.load(f)
        assert any(p["project_name"] == "테스트 과제명" for p in saved_projects)

        print("OK: test_refresh_read_seed_route")
    finally:
        read_seed.READ_DIR = original_read_dir
        history_store.HISTORY_PATH = original_history_path
        history_store.PROJECTS_PATH = original_projects_path
        shutil.rmtree(read_dir, ignore_errors=True)
        shutil.rmtree(data_dir, ignore_errors=True)


if __name__ == "__main__":
    test_refresh_read_seed_route()
    print("ALL PASSED")
