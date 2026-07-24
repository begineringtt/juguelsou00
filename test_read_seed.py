import os
import shutil
import tempfile

import openpyxl

from read_seed import scan_read_folder


class _ExplodingValue:
    """A cell value whose str() conversion always raises.

    Used to simulate a malformed cell that breaks `_cell_text`'s
    `str(value).strip()` call, so we can exercise the per-sheet
    try/except in `scan_read_folder` without needing a real, on-disk
    xlsx file that produces such a value (openpyxl only ever parses
    plain str/int/float/bool/datetime values out of real cells, so a
    value that explodes on str() can only be introduced in-memory).
    """

    def __str__(self):
        raise RuntimeError("boom: this cell value cannot be stringified")


def _build_fixture_workbook(path):
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "sheet1"
    ws1["A9"] = "업체명"
    ws1["D9"] = "주식회사 신안그린테크"
    ws1["L11"] = "청 구 인"
    ws1["O11"] = "김용직 주임연구원"
    ws1["A12"] = "내  용"
    ws1["D12"] = "고효율 과제 환경제어 자재 구매의 건"
    ws1["A13"] = "상세내용"
    ws1["B14"] = "해당 연구개발 과제 수행을 위한 자재를 구매 하오니 결재 승인 요청드립니다."
    ws1["B16"] = "중앙행정기관 : 농림축산식품부"
    ws1["B17"] = "전문기관 : 농림식품기술기획평가원"
    ws1["B18"] = "과제명 : 고효율 광원 및 지능형 광조절 시스템 탑재 모듈형 수직농장 모델 개발"
    ws1["B20"] = "연구재료비 집행의 건 (연구개발계획서 상 계상되어 있는 건임)"

    ws2 = wb.create_sheet("sheet2")
    ws2["A9"] = "업체명"
    ws2["D9"] = "(주)아이온이엔지"
    ws2["N11"] = "청 구 인"
    ws2["Q11"] = "이배훈 책임연구원"
    ws2["A12"] = "내  용"
    ws2["D12"] = "온실용 환경제어 계측 자재 구매 진행의 건"
    ws2["A13"] = "상세내용"
    ws2["C14"] = "해당 연구개발 수행을 위한 온실용 환경제어 계측 자재를 구매하고자 하오니 결재 승인 요청드립니다."
    ws2["B16"] = "중앙행정기관 : 농림축산식품부"
    ws2["B17"] = "전문기관 : 농림식품기술기획평가원"
    ws2["B18"] = "과제명 : 수확 후 전 과정 무인 자동화 시스템 개발 및 실증"
    ws2["B20"] = "연구재료비 집행의 건 (연구개발계획서 상 계상되어 있는 건임)"

    # Sheet 3 deliberately places every label 4 rows lower than sheet1/sheet2
    # (row 9 -> 13, row 11 -> 15, row 12 -> 16, row 13 -> 17, project block
    # 16-20 -> 20-24), with its own distinct values. This is the actual
    # regression check for label-based (row-independent) scanning: a
    # fixed-cell-offset implementation would miss every field on this sheet,
    # while label scanning should find them regardless of row position.
    ws3 = wb.create_sheet("sheet3_shifted")
    ws3["A13"] = "업체명"
    ws3["D13"] = "그린플러스 주식회사"
    ws3["L15"] = "청 구 인"
    ws3["O15"] = "박서연 선임연구원"
    ws3["A16"] = "내  용"
    ws3["D16"] = "스마트팜 통합관제 시스템 도입의 건"
    ws3["A17"] = "상세내용"
    ws3["B18"] = "본 과제 수행을 위한 스마트팜 통합관제 시스템 도입 건에 대해 결재를 요청드립니다."
    ws3["B20"] = "중앙행정기관 : 중소벤처기업부"
    ws3["B21"] = "전문기관 : 한국산업기술진흥원"
    ws3["B22"] = "과제명 : 지능형 스마트팜 통합관제 시스템 개발"
    ws3["B24"] = "연구재료비 집행의 건 (스마트팜 통합관제 예산 별도 계상)"

    ws4 = wb.create_sheet("제품등")
    ws4["A1"] = "관련 없는 시트"

    wb.save(path)


def _build_corrupted_file(path):
    with open(path, "wb") as f:
        f.write(b"not a real xlsx file")


def _build_exception_workbook():
    """Build (in memory, not saved to disk) a workbook whose first sheet
    contains a cell value that raises when `_scan_sheet` tries to read it,
    and whose second sheet is otherwise normal and valid.

    This is used to prove that `scan_read_folder`'s per-sheet
    `try/except Exception: continue` correctly isolates a single bad sheet
    without aborting the scan of the remaining sheets in that workbook.
    """
    wb = openpyxl.Workbook()

    ws_bad = wb.active
    ws_bad.title = "bad_sheet"
    ws_bad["A9"] = "업체명"
    # openpyxl validates cell values on assignment (`ws[...] = value`) and
    # rejects arbitrary objects, and any object that *is* accepted gets
    # flattened to a plain string once saved to and reloaded from a real
    # xlsx file. So to get a value that genuinely explodes on str() at scan
    # time, we bypass the setter and write directly to the underlying
    # attribute on an in-memory cell.
    ws_bad["D9"]._value = _ExplodingValue()

    ws_good = wb.create_sheet("good_sheet")
    ws_good["A9"] = "업체명"
    ws_good["D9"] = "정상시트 주식회사"
    ws_good["L11"] = "청 구 인"
    ws_good["O11"] = "최유진 연구원"
    ws_good["A12"] = "내  용"
    ws_good["D12"] = "정상 시트 자재 구매의 건"
    ws_good["A13"] = "상세내용"
    ws_good["B14"] = "정상 시트에서 추출되어야 하는 상세내용입니다."
    ws_good["B16"] = "중앙행정기관 : 산업통상자원부"
    ws_good["B17"] = "전문기관 : 한국산업기술평가관리원"
    ws_good["B18"] = "과제명 : 정상 시트 검증용 과제명"
    ws_good["B20"] = "연구재료비 집행의 건 (정상 시트 검증)"

    return wb


def test_scan_read_folder():
    tmp_dir = tempfile.mkdtemp()
    try:
        _build_fixture_workbook(os.path.join(tmp_dir, "fixture_a.xlsx"))

        skip_wb = openpyxl.Workbook()
        skip_wb.active["A9"] = "업체명"
        skip_wb.active["D9"] = "UNIQUE_SKIP_COMPANY_XYZ"
        skip_wb.save(os.path.join(tmp_dir, "~$fixture_a.xlsx"))

        _build_corrupted_file(os.path.join(tmp_dir, "broken.xlsx"))

        result = scan_read_folder(tmp_dir)

        assert result["history"]["company"] == [
            "주식회사 신안그린테크",
            "(주)아이온이엔지",
            "그린플러스 주식회사",
        ]
        assert result["history"]["requester"] == [
            "김용직 주임연구원",
            "이배훈 책임연구원",
            "박서연 선임연구원",
        ]
        assert result["history"]["title"] == [
            "고효율 과제 환경제어 자재 구매의 건",
            "온실용 환경제어 계측 자재 구매 진행의 건",
            "스마트팜 통합관제 시스템 도입의 건",
        ]
        assert result["history"]["detail"] == [
            "해당 연구개발 과제 수행을 위한 자재를 구매 하오니 결재 승인 요청드립니다.",
            "해당 연구개발 수행을 위한 온실용 환경제어 계측 자재를 구매하고자 하오니 결재 승인 요청드립니다.",
            "본 과제 수행을 위한 스마트팜 통합관제 시스템 도입 건에 대해 결재를 요청드립니다.",
        ]
        assert result["history"]["execution_note"] == [
            "연구재료비 집행의 건 (연구개발계획서 상 계상되어 있는 건임)",
            "연구재료비 집행의 건 (스마트팜 통합관제 예산 별도 계상)",
        ]
        assert "UNIQUE_SKIP_COMPANY_XYZ" not in result["history"]["company"]

        assert result["projects"] == [
            {
                "agency": "농림축산식품부",
                "org": "농림식품기술기획평가원",
                "project_name": "고효율 광원 및 지능형 광조절 시스템 탑재 모듈형 수직농장 모델 개발",
            },
            {
                "agency": "농림축산식품부",
                "org": "농림식품기술기획평가원",
                "project_name": "수확 후 전 과정 무인 자동화 시스템 개발 및 실증",
            },
            {
                "agency": "중소벤처기업부",
                "org": "한국산업기술진흥원",
                "project_name": "지능형 스마트팜 통합관제 시스템 개발",
            },
        ]
        print("OK: test_scan_read_folder")
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)


def test_scan_sheet_exception_does_not_abort_workbook():
    tmp_dir = tempfile.mkdtemp()
    try:
        # A real file must exist on disk for glob() to find it. Its actual
        # bytes are irrelevant: load_workbook is patched below to hand back
        # our hand-built in-memory workbook (containing the exploding cell)
        # for this exact path instead of parsing the file from disk.
        path = os.path.join(tmp_dir, "exception_fixture.xlsx")
        with open(path, "wb") as f:
            f.write(b"placeholder, replaced by patched load_workbook below")

        exception_wb = _build_exception_workbook()

        real_load_workbook = openpyxl.load_workbook

        def fake_load_workbook(p, *args, **kwargs):
            if os.path.abspath(p) == os.path.abspath(path):
                return exception_wb
            return real_load_workbook(p, *args, **kwargs)

        openpyxl.load_workbook = fake_load_workbook
        try:
            result = scan_read_folder(tmp_dir)
        finally:
            openpyxl.load_workbook = real_load_workbook

        # The bad sheet contributed nothing (its exception fired on the very
        # first field, "company", before anything from that sheet could be
        # recorded) -- only one company should have been extracted overall.
        assert len(result["history"]["company"]) == 1

        # ...while the sibling "good_sheet" in the SAME workbook must still
        # be fully scanned, proving the per-sheet try/except in
        # scan_read_folder isolated the failure to just the bad sheet.
        assert "정상시트 주식회사" in result["history"]["company"]
        assert "최유진 연구원" in result["history"]["requester"]
        assert "정상 시트 자재 구매의 건" in result["history"]["title"]
        assert "정상 시트에서 추출되어야 하는 상세내용입니다." in result["history"]["detail"]
        assert "연구재료비 집행의 건 (정상 시트 검증)" in result["history"]["execution_note"]
        assert any(
            p == {
                "agency": "산업통상자원부",
                "org": "한국산업기술평가관리원",
                "project_name": "정상 시트 검증용 과제명",
            }
            for p in result["projects"]
        )
        print("OK: test_scan_sheet_exception_does_not_abort_workbook")
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)


def test_seal_placeholder_is_not_treated_as_requester():
    tmp_dir = tempfile.mkdtemp()
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A11"] = "지출일"
        ws["N11"] = "청 구 인"
        ws["AC11"] = "(인)"
        wb.save(os.path.join(tmp_dir, "seal_only.xlsx"))

        result = scan_read_folder(tmp_dir)

        assert result["history"]["requester"] == []
        print("OK: test_seal_placeholder_is_not_treated_as_requester")
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)


if __name__ == "__main__":
    test_scan_read_folder()
    test_scan_sheet_exception_does_not_abort_workbook()
    test_seal_placeholder_is_not_treated_as_requester()
    print("ALL PASSED")
