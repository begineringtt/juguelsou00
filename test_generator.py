import openpyxl
from generator import build_expense_report, ITEM_HEADER_ROW, LAST_USABLE_ROW

BASE_DATA = {
    "company": "주식회사 테스트",
    "propose_date": "2026-07-22",
    "spend_date": "2026-07-22",
    "department": "그린연구소",
    "requester": "홍길동 연구원",
    "title": "테스트 구매의 건",
    "detail": "테스트 상세내용입니다.",
    "agency": "농림축산식품부",
    "org": "농림식품기술기획평가원",
    "project_name": "테스트 과제명",
    "execution_note": "연구재료비 집행의 건",
}


def make_items(n):
    return [
        {"name": f"품목{i+1}", "spec": f"규격{i+1}", "unit": "EA", "qty": i + 1, "price": 1000 * (i + 1)}
        for i in range(n)
    ]


def inspect(n):
    data = dict(BASE_DATA)
    data["items"] = make_items(n)
    buf = build_expense_report(data)
    wb = openpyxl.load_workbook(buf)
    ws = wb.worksheets[0]
    print(f"\n===== n={n} items =====")
    print("dims:", ws.dimensions)
    print("merges around item area:")
    for mc in sorted(ws.merged_cells.ranges, key=lambda m: (m.min_row, m.min_col)):
        if ITEM_HEADER_ROW <= mc.min_row <= LAST_USABLE_ROW + 1:
            print(" ", mc)
    print("formulas:")
    for r in range(ITEM_HEADER_ROW, LAST_USABLE_ROW + 1):
        vals = []
        for col in ["B", "G", "K", "N", "Q", "U", "Y"]:
            v = ws[f"{col}{r}"].value
            if v is not None:
                vals.append(f"{col}{r}={v!r}")
        if vals:
            print(" ", " | ".join(vals))
    print("S5:", ws["S5"].value, "D5:", ws["D5"].value)
    # 고정 위치 결재란/각주가 그대로 남아있는지 확인
    for coord in ["A31", "M31"]:
        print(f"{coord}:", ws[coord].value)


for n in [1, 2, 3, 5, 9]:
    inspect(n)
