from openpyxl.utils import get_column_letter

from generator import _compute_column_layout, _infer_flags


def _letters(layout, key):
    start, end, _label = layout[key]
    return get_column_letter(start), get_column_letter(end)


def test_all_columns_matches_original_layout():
    flags = {"use_spec": True, "use_unit": True, "use_qty": True, "use_price": True}
    layout = _compute_column_layout(flags)
    assert _letters(layout, "name") == ("B", "F")
    assert _letters(layout, "spec") == ("G", "J")
    assert _letters(layout, "unit") == ("K", "M")
    assert _letters(layout, "qty") == ("N", "P")
    assert _letters(layout, "price") == ("Q", "T")
    assert _letters(layout, "supply") == ("U", "X")
    assert _letters(layout, "vat") == ("Y", "AB")
    print("OK: test_all_columns_matches_original_layout")


def test_spec_dropped_fills_27_columns():
    flags = {"use_spec": False, "use_unit": True, "use_qty": True, "use_price": True}
    layout = _compute_column_layout(flags)
    assert "spec" not in layout
    assert set(layout.keys()) == {"name", "unit", "qty", "price", "supply", "vat"}
    ranges = sorted((v[0], v[1]) for v in layout.values())
    assert ranges[0][0] == 2
    assert ranges[-1][1] == 28
    for i in range(len(ranges) - 1):
        assert ranges[i][1] + 1 == ranges[i + 1][0]
    print("OK: test_spec_dropped_fills_27_columns")


def test_all_optional_columns_dropped():
    flags = {"use_spec": False, "use_unit": False, "use_qty": False, "use_price": False}
    layout = _compute_column_layout(flags)
    assert set(layout.keys()) == {"name", "supply", "vat"}
    assert _letters(layout, "name") == ("B", "L")
    assert _letters(layout, "supply") == ("M", "T")
    assert _letters(layout, "vat") == ("U", "AB")
    print("OK: test_all_optional_columns_dropped")


def test_infer_flags_from_items():
    items = [{"name": "a", "spec": "x", "qty": 1, "price": 100}]
    flags = _infer_flags(items)
    assert flags == {"use_spec": True, "use_unit": False, "use_qty": True, "use_price": True}
    print("OK: test_infer_flags_from_items")


if __name__ == "__main__":
    test_all_columns_matches_original_layout()
    test_spec_dropped_fills_27_columns()
    test_all_optional_columns_dropped()
    test_infer_flags_from_items()
    print("ALL PASSED")
