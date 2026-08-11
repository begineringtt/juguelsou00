import openpyxl
from openpyxl.utils import get_column_letter

from generator import (
    ITEM_HEADER_ROW, TABLE_START_COL, TABLE_END_COL, FIRST_ITEM_ROW, BASE_ITEM_ROWS,
    FOOTER_BLOCK_START_ROW, FOOTER_BLOCK_MAX_COL, _compute_column_layout,
)
from generator import build_expense_report

BASE_DATA = {
    "company": "주식회사 테스트",
    "propose_date": "2026-07-22",
    "spend_date": "2026-07-22",
    "department": "그린연구소",
    "requester": "홍길동 연구원",
    "title": "테스트 구매의 건",
    "detail": "테스트 상세내용입니다.",
    "agency": "농림축산식품부",
    "org": "농림식품기술기획평가원",
    "project_name": "테스트 과제명",
    "execution_note": "연구재료비 집행의 건",
}


def _header_merges(ws):
    return sorted(
        (m.min_col, m.max_col)
        for m in ws.merged_cells.ranges
        if m.min_row == ITEM_HEADER_ROW and m.max_row == ITEM_HEADER_ROW
        and m.min_col >= TABLE_START_COL and m.max_col <= TABLE_END_COL
    )


def _header_labels(ws):
    labels = {}
    for col in range(TABLE_START_COL, TABLE_END_COL + 1):
        v = ws.cell(row=ITEM_HEADER_ROW, column=col).value
        if v:
            labels[get_column_letter(col)] = v
    return labels


def test_all_columns_regression():
    data = dict(BASE_DATA)
    data["items"] = [{"name": "품목1", "spec": "규격1", "unit": "EA", "qty": 2, "price": 1000}]
    buf = build_expense_report(data)
    wb = openpyxl.load_workbook(buf)
    ws = wb.worksheets[0]

    assert _header_labels(ws) == {
        "B": "품목", "G": "규격", "K": "단위", "N": "수량", "Q": "단가", "U": "공급가", "Y": "부가세",
    }
    assert _header_merges(ws) == sorted([(2, 6), (7, 10), (11, 13), (14, 16), (17, 20), (21, 24), (25, 28)])
    print("OK: test_all_columns_regression")


def test_spec_dropped_compacts_header():
    data = dict(BASE_DATA)
    data["items"] = [{"name": "품목1", "unit": "EA", "qty": 2, "price": 1000}]
    buf = build_expense_report(data)
    wb = openpyxl.load_workbook(buf)
    ws = wb.worksheets[0]

    labels = _header_labels(ws)
    assert "규격" not in labels.values()
    assert set(labels.values()) == {"품목", "단위", "수량", "단가", "공급가", "부가세"}

    merges = _header_merges(ws)
    assert merges[0][0] == 2
    assert merges[-1][1] == 28
    total_span = sum(end - start + 1 for start, end in merges)
    assert total_span == 27
    print("OK: test_spec_dropped_compacts_header")


def test_total_row_boundary_matches_supply_start():
    data = dict(BASE_DATA)
    data["items"] = [{"name": "품목1", "unit": "EA", "qty": 2, "price": 1000}]
    flags = {"use_spec": False, "use_unit": True, "use_qty": True, "use_price": True}
    layout = _compute_column_layout(flags)
    supply_start = layout["supply"][0]

    buf = build_expense_report(data)
    wb = openpyxl.load_workbook(buf)
    ws = wb.worksheets[0]

    total_row = None
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == "합계 금액":
                total_row = cell.row
                break
        if total_row:
            break
    assert total_row is not None

    row_merges = [
        (m.min_col, m.max_col) for m in ws.merged_cells.ranges
        if m.min_row == total_row and m.max_row == total_row
    ]
    label_merge = next(mc for mc in row_merges if mc[0] == TABLE_START_COL)
    amount_merge = next(mc for mc in row_merges if mc[0] == supply_start)
    assert label_merge[1] == supply_start - 1
    assert amount_merge[1] == TABLE_END_COL
    print("OK: test_total_row_boundary_matches_supply_start")


def test_supply_direct_entry_when_price_off():
    data = dict(BASE_DATA)
    data["items"] = [{"name": "품목1", "spec": "규격1", "unit": "EA", "supply": 12345}]
    buf = build_expense_report(data)
    wb = openpyxl.load_workbook(buf)
    ws = wb.worksheets[0]
    layout = _compute_column_layout({"use_spec": True, "use_unit": True, "use_qty": False, "use_price": False})
    supply_letter = get_column_letter(layout["supply"][0])
    assert ws[f"{supply_letter}{FIRST_ITEM_ROW}"].value == 12345
    print("OK: test_supply_direct_entry_when_price_off")


def test_qty_price_supply_cells_populated():
    data = dict(BASE_DATA)
    data["items"] = [{"name": "품목1", "spec": "규격1", "unit": "EA", "qty": 2, "price": 1000}]
    buf = build_expense_report(data)
    wb = openpyxl.load_workbook(buf)
    ws = wb.worksheets[0]

    layout = _compute_column_layout({"use_spec": True, "use_unit": True, "use_qty": True, "use_price": True})
    qty_letter = get_column_letter(layout["qty"][0])
    price_letter = get_column_letter(layout["price"][0])
    supply_letter = get_column_letter(layout["supply"][0])

    assert ws[f"{qty_letter}{FIRST_ITEM_ROW}"].value == 2
    assert ws[f"{price_letter}{FIRST_ITEM_ROW}"].value == 1000
    assert ws[f"{supply_letter}{FIRST_ITEM_ROW}"].value == f"={qty_letter}{FIRST_ITEM_ROW}*{price_letter}{FIRST_ITEM_ROW}"
    print("OK: test_qty_price_supply_cells_populated")


def test_base_item_count_leaves_spacer_row_before_footer():
    data = dict(BASE_DATA)
    data["items"] = [{"name": f"품목{i+1}", "unit": "EA", "qty": 1, "price": 100} for i in range(BASE_ITEM_ROWS)]
    buf = build_expense_report(data)
    wb = openpyxl.load_workbook(buf)
    ws = wb.worksheets[0]

    total_row = FIRST_ITEM_ROW + BASE_ITEM_ROWS + 1  # +1 for 이하 여백
    assert ws.cell(row=total_row, column=2).value == "합계 금액"

    spacer_row = total_row + 1
    assert ws.cell(row=spacer_row, column=2).value is None
    assert ws.cell(row=spacer_row, column=1).border.left.style == "medium"

    footer_start = spacer_row + 1
    assert footer_start == FOOTER_BLOCK_START_ROW + 1
    assert ws.cell(row=footer_start, column=17).value == "재\n무\n부\n서"
    footer_row = footer_start + 3
    assert ws.cell(row=footer_row, column=1).value == "[GP-A-001]"
    assert ws.cell(row=footer_row, column=13).value == "주식회사 그린플러스"
    print("OK: test_base_item_count_leaves_spacer_row_before_footer")


def test_many_items_inserts_rows_and_pushes_footer_block_down():
    extra = 6
    n = BASE_ITEM_ROWS + extra
    data = dict(BASE_DATA)
    data["items"] = [{"name": f"품목{i+1}", "unit": "EA", "qty": 1, "price": 100} for i in range(n)]
    buf = build_expense_report(data)
    wb = openpyxl.load_workbook(buf)
    ws = wb.worksheets[0]

    last_item_row = FIRST_ITEM_ROW + n - 1
    assert ws.cell(row=last_item_row, column=2).value == f"품목{n}"
    assert ws.row_dimensions[last_item_row].height == 21.75

    blank_row = last_item_row + 1
    total_row = blank_row + 1
    assert ws.cell(row=blank_row, column=2).value == "이하 여백"
    assert ws.cell(row=total_row, column=2).value == "합계 금액"

    spacer_row = total_row + 1
    assert ws.cell(row=spacer_row, column=2).value is None

    new_footer_start = spacer_row + 1
    assert new_footer_start == FOOTER_BLOCK_START_ROW + extra + 1
    assert ws.cell(row=new_footer_start, column=17).value == "재\n무\n부\n서"
    assert ws.cell(row=new_footer_start, column=18).value == "담당"

    footer_row = new_footer_start + 3
    assert ws.cell(row=footer_row, column=1).value == "[GP-A-001]"
    assert ws.cell(row=footer_row, column=13).value == "주식회사 그린플러스"

    # 재무부서 결재란/각주 병합도 새 위치로 그대로 옮겨졌는지 확인
    merges = {(m.min_row, m.max_row, m.min_col, m.max_col) for m in ws.merged_cells.ranges}
    assert (new_footer_start, new_footer_start + 2, 17, 17) in merges  # Q열 세로 병합 (재무부서)
    assert (footer_row, footer_row, 1, 12) in merges  # [GP-A-001]
    assert (footer_row, footer_row, 13, 29) in merges  # 주식회사 그린플러스

    # 페이지 나누기 미리보기(인쇄 영역)도 새 마지막 행까지 늘어났는지 확인
    assert ws.print_area == f"'{ws.title}'!$A$1:$AC${footer_row}"
    print("OK: test_many_items_inserts_rows_and_pushes_footer_block_down")


def test_many_items_keeps_outer_frame_border_consistent():
    """품목이 넘쳐서 원본 캔버스(11~27행) 밖으로 나간 행도, 안쪽 행과 똑같이
    A열/AC열 바깥 테두리(medium)가 이어지는지 확인한다.
    """
    n = BASE_ITEM_ROWS + 6
    data = dict(BASE_DATA)
    data["items"] = [{"name": f"품목{i+1}", "unit": "EA", "qty": 1, "price": 100} for i in range(n)]
    buf = build_expense_report(data)
    wb = openpyxl.load_workbook(buf)
    ws = wb.worksheets[0]

    inside_canvas_row = FIRST_ITEM_ROW + 1          # 원본 11~27행 범위 안
    overflow_row = FIRST_ITEM_ROW + BASE_ITEM_ROWS + 2  # 원본 범위를 넘어간 품목 행
    for row in (ITEM_HEADER_ROW, inside_canvas_row, overflow_row):
        left = ws.cell(row=row, column=1).border.left
        right = ws.cell(row=row, column=FOOTER_BLOCK_MAX_COL).border.right
        assert left and left.style == "medium", f"row {row} missing left frame border"
        assert right and right.style == "medium", f"row {row} missing right frame border"
    print("OK: test_many_items_keeps_outer_frame_border_consistent")


if __name__ == "__main__":
    test_all_columns_regression()
    test_spec_dropped_compacts_header()
    test_total_row_boundary_matches_supply_start()
    test_supply_direct_entry_when_price_off()
    test_qty_price_supply_cells_populated()
    test_base_item_count_leaves_spacer_row_before_footer()
    test_many_items_inserts_rows_and_pushes_footer_block_down()
    test_many_items_keeps_outer_frame_border_consistent()
    print("ALL PASSED")
