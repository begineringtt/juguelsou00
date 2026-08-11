"""지출결의서(支出決議書) 엑셀 파일 자동 생성 로직.

template_files/base_template.xlsx 를 원본으로 하여, 사용자가 입력한 값으로
채운 새 엑셀 파일을 만들어 낸다. 품목 개수는 1개 이상 자유롭게 늘어날 수 있고,
기본 칸 수(9개, LAST_USABLE_ROW 기준)를 넘으면 그만큼 행을 더 만들어서
재무부서 결재란/회사명 각주(고정 블록)를 그 아래로 밀어낸다.

rev.2 양식은 품목 표/과제 정보 영역(11~27행)이 완전히 빈 캔버스로 제공되어
서식을 캡처할 참조 행이 없다. 그래서 폰트/테두리를 코드에서 직접 구성해서
매번 새로 그리는 방식으로 구현한다. 결재란/각주 블록(28~31행)은 반대로 이미
서식이 잡혀 있으므로, 매번 그 서식을 통째로 캡처해뒀다가 품목 수에 맞는
위치에 다시 그려 넣는다.
"""

import copy
import datetime
import io
import os

import openpyxl
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

from paths import bundle_dir

BASE_DIR = bundle_dir()
TEMPLATE_PATH = os.path.join(BASE_DIR, "template_files", "base_template.xlsx")

SHEET_NAME = "지출결의서"

FONT_NAME = "굴림"
FONT_REGULAR = Font(name=FONT_NAME, size=10)
FONT_BOLD = Font(name=FONT_NAME, size=10, bold=True)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_LEFT = Alignment(horizontal="left", vertical="center")

# 원본 템플릿이 수량/단가/공급가/부가세 칸(11~27행 O~AA열)에 미리 적용해 둔 회계
# 표시 형식. 그 범위 밖(품목이 많아 새로 생긴 행)의 숫자 칸에도 명시적으로
# 적용해야 "일반" 형식으로 보이지 않는다.
ACCOUNTING_NUMBER_FORMAT = '_-* #,##0_-;\\-* #,##0_-;_-* "-"_-;_-@_-'

_THIN_SIDE = Side(style="thin")
_NO_SIDE = Side(style=None)

# base_template.xlsx (rev.2) 기준 고정 좌표
PROJECT_INFO_ROWS = {
    "agency": 11,
    "org": 12,
    "project_name": 13,
    "execution_note": 14,
}
PROJECT_INFO_START_COL = 2  # B
PROJECT_INFO_END_COL = 29   # AC

ITEM_HEADER_ROW = 16
FIRST_ITEM_ROW = 17
LAST_USABLE_ROW = 27  # 원본 템플릿에서 결재란 앞까지 비어 있는 마지막 행

TABLE_START_COL = 2   # B
TABLE_END_COL = 28    # AB
TABLE_WIDTH = TABLE_END_COL - TABLE_START_COL + 1  # 27

# 품목 헤더(1) + 이하 여백(1) + 합계 금액(1) 을 제외한 나머지가 기본 품목 칸 수.
# 이보다 품목이 많으면 그만큼 행을 더 만들고 아래 결재란/각주 블록을 밀어낸다.
BASE_ITEM_ROWS = LAST_USABLE_ROW - FIRST_ITEM_ROW - 1  # 9

# 재무부서 결재란 + 회사명 각주(고정 서식 블록). 원본 템플릿 기준 28~31행.
FOOTER_BLOCK_START_ROW = LAST_USABLE_ROW + 1  # 28
FOOTER_BLOCK_ROW_COUNT = 4
FOOTER_BLOCK_MAX_COL = PROJECT_INFO_END_COL   # AC, 블록이 쓰는 가장 오른쪽 열

# 11~27행(빈 캔버스) 원본 행 높이. 이 범위를 벗어나는 새 행(품목이 많아서 늘어난
# 행, 합계 금액 아래 여백 행)에도 그대로 적용해서 서식이 어긋나지 않게 한다.
ITEM_ROW_HEIGHT = 21.75

ITEM_FIELDS = [
    ("name", "품목", 4, None),
    ("spec", "규격", 3, "use_spec"),
    ("unit", "단위", 2, "use_unit"),
    ("qty", "수량", 2, "use_qty"),
    ("price", "단가", 3, "use_price"),
    ("supply", "공급가", 3, None),
    ("vat", "부가세", 3, None),
]


def _largest_remainder_allocation(weights, total):
    weight_sum = sum(weights)
    raw = [w / weight_sum * total for w in weights]
    floors = [int(x) for x in raw]
    remainder = total - sum(floors)
    order = sorted(range(len(weights)), key=lambda i: raw[i] - floors[i], reverse=True)
    for i in order[:remainder]:
        floors[i] += 1
    return floors


def _infer_flags(items):
    return {
        "use_spec": any("spec" in item for item in items),
        "use_unit": any("unit" in item for item in items),
        "use_qty": any("qty" in item for item in items),
        "use_price": any("price" in item for item in items),
    }


def _compute_column_layout(flags):
    active = [f for f in ITEM_FIELDS if f[3] is None or flags.get(f[3])]
    weights = [f[2] for f in active]
    spans = _largest_remainder_allocation(weights, TABLE_WIDTH)

    layout = {}
    col = TABLE_START_COL
    for (key, label, _weight, _flag), span in zip(active, spans):
        layout[key] = (col, col + span - 1, label)
        col += span
    return layout


def _col_letter(layout, key):
    return get_column_letter(layout[key][0])


def _style_span(ws, row, start_col, end_col, font, alignment=ALIGN_CENTER):
    """병합 구간(start_col~end_col)에 테두리(바깥쪽 thin)와 폰트/정렬을 적용한다."""
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.border = Border(
            left=_THIN_SIDE if col == start_col else _NO_SIDE,
            right=_THIN_SIDE if col == end_col else _NO_SIDE,
            top=_THIN_SIDE,
            bottom=_THIN_SIDE,
        )
        cell.font = font
        cell.alignment = alignment


def _add_item_row_merges(ws, row, layout):
    for start, end, _label in layout.values():
        if end > start:
            ws.merge_cells(start_row=row, start_column=start, end_row=row, end_column=end)


def _apply_outer_frame(ws, row):
    """품목 표 영역을 감싸는 바깥 테두리(A열 왼쪽/AC열 오른쪽, medium)를 적용한다.

    원본 템플릿은 11~27행에만 이 테두리가 미리 그려져 있어서, 품목이 많아 그
    범위를 벗어난 새 행은 그냥 두면 테두리가 끊긴다. 그래서 품목 표와 관련된
    모든 행(헤더/품목/이하여백/합계금액/여유행)에 매번 명시적으로 적용한다.
    """
    ws.cell(row=row, column=1).border = Border(left=Side(style="medium"))
    ws.cell(row=row, column=FOOTER_BLOCK_MAX_COL).border = Border(right=Side(style="medium"))


def _rebuild_header_row(ws, layout):
    ws.row_dimensions[ITEM_HEADER_ROW].height = ITEM_ROW_HEIGHT
    _apply_outer_frame(ws, ITEM_HEADER_ROW)
    for start, end, label in layout.values():
        if end > start:
            ws.merge_cells(start_row=ITEM_HEADER_ROW, start_column=start,
                            end_row=ITEM_HEADER_ROW, end_column=end)
        ws.cell(row=ITEM_HEADER_ROW, column=start).value = label
        _style_span(ws, ITEM_HEADER_ROW, start, end, FONT_BOLD)


def _write_item_row(ws, row, layout, item, supply_letter):
    ws.row_dimensions[row].height = ITEM_ROW_HEIGHT
    _apply_outer_frame(ws, row)
    ws[f"{_col_letter(layout, 'name')}{row}"] = item["name"]
    if "spec" in layout and item.get("spec"):
        ws[f"{_col_letter(layout, 'spec')}{row}"] = item["spec"]
    if "unit" in layout and item.get("unit"):
        ws[f"{_col_letter(layout, 'unit')}{row}"] = item["unit"]

    use_qty = "qty" in layout and item.get("qty") is not None
    use_price = "price" in layout and item.get("price") is not None
    if use_qty:
        qty_cell = ws[f"{_col_letter(layout, 'qty')}{row}"]
        qty_cell.value = item["qty"]
        qty_cell.number_format = ACCOUNTING_NUMBER_FORMAT
    if use_price:
        price_letter = _col_letter(layout, "price")
        price_cell = ws[f"{price_letter}{row}"]
        price_cell.value = item["price"]
        price_cell.number_format = ACCOUNTING_NUMBER_FORMAT
        if use_qty:
            qty_letter = _col_letter(layout, "qty")
            ws[f"{supply_letter}{row}"] = f"={qty_letter}{row}*{price_letter}{row}"
        else:
            ws[f"{supply_letter}{row}"] = f"={price_letter}{row}"
    else:
        ws[f"{supply_letter}{row}"] = item.get("supply", 0)
    supply_cell = ws[f"{supply_letter}{row}"]
    supply_cell.number_format = ACCOUNTING_NUMBER_FORMAT
    vat_cell = ws[f"{_col_letter(layout, 'vat')}{row}"]
    vat_cell.value = f"={supply_letter}{row}/10"
    vat_cell.number_format = ACCOUNTING_NUMBER_FORMAT

    _add_item_row_merges(ws, row, layout)
    for start, end, _label in layout.values():
        _style_span(ws, row, start, end, FONT_REGULAR)


def _capture_footer_block(ws):
    """재무부서 결재란 + 회사명 각주 블록(FOOTER_BLOCK_START_ROW 기준 4행)의
    값/서식/병합 정보를 통째로 캡처한다. 품목이 기본 칸 수를 넘어가면 이 블록을
    지웠다가 더 아래 위치에 다시 그려 넣기 위함이다.
    """
    rows = []
    for offset in range(FOOTER_BLOCK_ROW_COUNT):
        row = FOOTER_BLOCK_START_ROW + offset
        cells = []
        for col in range(1, FOOTER_BLOCK_MAX_COL + 1):
            c = ws.cell(row=row, column=col)
            cells.append({
                "value": c.value,
                "font": copy.copy(c.font),
                "border": copy.copy(c.border),
                "fill": copy.copy(c.fill),
                "alignment": copy.copy(c.alignment),
                "number_format": c.number_format,
            })
        rows.append({"height": ws.row_dimensions[row].height, "cells": cells})

    merges = [
        (m.min_row - FOOTER_BLOCK_START_ROW, m.max_row - FOOTER_BLOCK_START_ROW, m.min_col, m.max_col)
        for m in ws.merged_cells.ranges
        if m.min_row >= FOOTER_BLOCK_START_ROW
    ]
    return {"rows": rows, "merges": merges}


def _clear_footer_block(ws):
    for m in [m for m in ws.merged_cells.ranges if m.min_row >= FOOTER_BLOCK_START_ROW]:
        ws.unmerge_cells(str(m))
    ws.delete_rows(FOOTER_BLOCK_START_ROW, FOOTER_BLOCK_ROW_COUNT)


def _restore_footer_block(ws, block, new_start_row):
    for offset, row_data in enumerate(block["rows"]):
        row = new_start_row + offset
        ws.row_dimensions[row].height = row_data["height"]
        for col_idx, style in enumerate(row_data["cells"], start=1):
            c = ws.cell(row=row, column=col_idx)
            c.value = style["value"]
            c.font = style["font"]
            c.border = style["border"]
            c.fill = style["fill"]
            c.alignment = style["alignment"]
            c.number_format = style["number_format"]

    for row_off_start, row_off_end, min_col, max_col in block["merges"]:
        ws.merge_cells(
            start_row=new_start_row + row_off_start,
            end_row=new_start_row + row_off_end,
            start_column=min_col,
            end_column=max_col,
        )


def _rebuild_item_section(ws, items):
    """품목 표(헤더~합계 금액)를 11행 이후 빈 캔버스에 새로 그린다.

    품목이 BASE_ITEM_ROWS(기본 9개)를 넘어가면 그만큼 행을 더 만들고, 아래에
    고정되어 있던 재무부서 결재란/각주 블록을 그 뒤로 밀어서 다시 그린다.

    반환값: total_row (합계 금액 행 번호)
    """
    footer_block = _capture_footer_block(ws)
    _clear_footer_block(ws)

    flags = _infer_flags(items)
    layout = _compute_column_layout(flags)
    supply_letter = _col_letter(layout, "supply")

    _rebuild_header_row(ws, layout)

    row = FIRST_ITEM_ROW
    for item in items:
        _write_item_row(ws, row, layout, item, supply_letter)
        row += 1

    table_start_letter = get_column_letter(TABLE_START_COL)

    blank_row = row
    table_end_letter = get_column_letter(TABLE_END_COL)
    ws.row_dimensions[blank_row].height = ITEM_ROW_HEIGHT
    _apply_outer_frame(ws, blank_row)
    ws.merge_cells(f"{table_start_letter}{blank_row}:{table_end_letter}{blank_row}")
    ws[f"{table_start_letter}{blank_row}"] = "이하 여백"
    _style_span(ws, blank_row, TABLE_START_COL, TABLE_END_COL, FONT_REGULAR)
    row += 1

    total_row = row
    ws.row_dimensions[total_row].height = ITEM_ROW_HEIGHT
    _apply_outer_frame(ws, total_row)
    ws[f"{table_start_letter}{total_row}"] = "합계 금액"
    total_amount_cell = ws[f"{supply_letter}{total_row}"]
    total_amount_cell.value = f"=SUM({supply_letter}{FIRST_ITEM_ROW}:{table_end_letter}{blank_row})"
    total_amount_cell.number_format = ACCOUNTING_NUMBER_FORMAT
    supply_start_col = layout["supply"][0]
    label_end_letter = get_column_letter(supply_start_col - 1)
    ws.merge_cells(f"{table_start_letter}{total_row}:{label_end_letter}{total_row}")
    ws.merge_cells(f"{supply_letter}{total_row}:{table_end_letter}{total_row}")
    _style_span(ws, total_row, TABLE_START_COL, supply_start_col - 1, FONT_BOLD)
    _style_span(ws, total_row, supply_start_col, TABLE_END_COL, FONT_BOLD)

    ws["S5"] = f"={supply_letter}{total_row}"

    # 합계 금액 행과 결재란 사이에 여유 있게 빈 행을 하나 두되, 문서 전체를
    # 감싸는 바깥 테두리(A/AC 열)는 끊기지 않도록 이어준다.
    spacer_row = total_row + 1
    ws.row_dimensions[spacer_row].height = ITEM_ROW_HEIGHT
    _apply_outer_frame(ws, spacer_row)

    footer_start_row = spacer_row + 1
    _restore_footer_block(ws, footer_block, footer_start_row)

    last_row = footer_start_row + FOOTER_BLOCK_ROW_COUNT - 1
    ws.print_area = f"A1:{get_column_letter(FOOTER_BLOCK_MAX_COL)}{last_row}"

    return total_row


def _write_project_info_line(ws, row, text):
    end_letter = get_column_letter(PROJECT_INFO_END_COL)
    ws.merge_cells(f"B{row}:{end_letter}{row}")
    cell = ws.cell(row=row, column=PROJECT_INFO_START_COL)
    cell.value = text
    cell.font = FONT_BOLD
    cell.alignment = ALIGN_LEFT


def _to_excel_date(value):
    if not value:
        return None
    if isinstance(value, (datetime.date, datetime.datetime)):
        return value
    return datetime.datetime.strptime(value, "%Y-%m-%d").date()


def build_expense_report(data):
    """입력 데이터(dict)를 받아 완성된 워크북을 BytesIO 로 반환한다.

    data 예시는 모듈 하단 __main__ 참고.
    """
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ws = wb.worksheets[0]
    ws.title = SHEET_NAME

    ws["D6"] = data.get("company", "")
    ws["Q6"] = data.get("doc_number") or None

    propose_date = _to_excel_date(data.get("propose_date"))
    if propose_date:
        ws["D7"] = propose_date
        ws["D7"].number_format = "yyyy-mm-dd"

    spend_date = _to_excel_date(data.get("spend_date"))
    if spend_date:
        ws["D8"] = spend_date
        ws["D8"].number_format = "yyyy-mm-dd"

    ws["Q7"] = data.get("department", "그린연구소")
    ws["Q8"] = data.get("requester", "")
    ws["D9"] = data.get("title", "")
    ws["D10"] = data.get("detail", "")

    agency = data.get("agency", "")
    org = data.get("org", "")
    project_name = data.get("project_name", "")
    _write_project_info_line(ws, PROJECT_INFO_ROWS["agency"], f"중앙행정기관 : {agency}" if agency else "")
    _write_project_info_line(ws, PROJECT_INFO_ROWS["org"], f"전문기관 : {org}" if org else "")
    _write_project_info_line(ws, PROJECT_INFO_ROWS["project_name"], f"과제명 : {project_name}" if project_name else "")
    _write_project_info_line(ws, PROJECT_INFO_ROWS["execution_note"], data.get("execution_note", ""))

    items = data.get("items") or []
    if not items:
        raise ValueError("품목이 최소 1개 이상 필요합니다.")
    _rebuild_item_section(ws, items)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def suggest_filename(data):
    company = (data.get("company") or "업체명미상").replace(" ", "")
    date_str = (data.get("propose_date") or "").replace("-", "")
    if not date_str:
        date_str = "00000000"
    return f"지출결의서_{company}_{date_str}.xlsx"
